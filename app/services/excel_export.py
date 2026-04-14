from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font


def _detect_bank_format(transactions: list[dict[str, Any]]) -> str:
    """Detect which bank format to use based on transaction data fields."""
    for row in transactions:
        bank_id = row.get("bank_id", "")
        if bank_id == "capitec_personal":
            return "capitec_personal"
        if bank_id == "capitec":
            return "capitec_business"
    if any(row.get("transaction_date") is not None or row.get("reference") is not None for row in transactions):
        return "capitec_business"
    return "default"


def build_excel_bytes(transactions: list[dict[str, Any]]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    fmt = _detect_bank_format(transactions)

    if fmt == "capitec_personal":
        headers = ["Date", "Description", "Category", "Money In", "Money Out", "Fee", "Balance"]
    elif fmt == "capitec_business":
        headers = ["Post Date", "Trans. Date", "Description", "Reference", "Fees", "Amount", "Balance"]
    else:
        headers = ["Date", "Description", "Amount", "Balance", "Accrued Bank Charges"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in transactions:
        if fmt == "capitec_personal":
            ws.append([
                row.get("date"),
                row.get("description"),
                row.get("category"),
                row.get("money_in"),
                row.get("money_out"),
                row.get("charges"),
                row.get("balance"),
            ])
        elif fmt == "capitec_business":
            ws.append([
                row.get("post_date") or row.get("date"),
                row.get("transaction_date"),
                row.get("description"),
                row.get("reference"),
                row.get("charges"),
                row.get("amount"),
                row.get("balance"),
            ])
        else:
            ws.append([
                row.get("date"),
                row.get("description"),
                row.get("amount"),
                row.get("balance"),
                row.get("charges"),
            ])

    ws.freeze_panes = "A2"
    if fmt == "capitec_personal":
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 36
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 10
        ws.column_dimensions["G"].width = 14
        numeric_cols = ("D", "E", "F", "G")
    elif fmt == "capitec_business":
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 36
        ws.column_dimensions["D"].width = 44
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 14
        ws.column_dimensions["G"].width = 14
        numeric_cols = ("E", "F", "G")
    else:
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 60
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 20
        numeric_cols = ("C", "D", "E")

    for col in numeric_cols:
        for cell in ws[col][1:]:
            cell.number_format = "#,##0.00"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
